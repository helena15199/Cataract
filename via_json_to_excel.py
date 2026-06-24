"""Convert a VIA (VGG Image Annotator) JSON annotation file to Excel.

Usage:
    python via_json_to_excel.py annotations.json
    python via_json_to_excel.py annotations.json --out results.xlsx
"""

import argparse
import json
import pathlib


def seconds_to_hms(seconds: float) -> str:
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:06.3f}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("json_file", help="Path to VIA JSON annotation file")
    parser.add_argument("--out", default=None, help="Output Excel file path (default: same name as input)")
    args = parser.parse_args()

    json_path = pathlib.Path(args.json_file)
    out_path  = pathlib.Path(args.out) if args.out else json_path.with_suffix(".xlsx")

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    # Extract attribute names (e.g. {"1": "Activity", "2": "Object"})
    attributes = {
        aid: ainfo["aname"]
        for aid, ainfo in data.get("attribute", {}).items()
    }

    # Extract file names
    files = {
        str(finfo["fid"]): finfo["fname"]
        for finfo in data.get("file", {}).values()
    }

    # Build rows
    rows = []
    for ann in data.get("metadata", {}).values():
        vid      = ann.get("vid", "")
        fname    = files.get(str(vid), vid)
        z        = ann.get("z", [])
        start    = float(z[0]) if len(z) > 0 else 0.0
        end      = float(z[1]) if len(z) > 1 else start
        duration = end - start
        av       = ann.get("av", {})

        row = {
            "File":           fname,
            "Start (s)":      round(start, 3),
            "End (s)":        round(end, 3),
            "Duration (s)":   round(duration, 3),
            "Start (HH:MM:SS)":   seconds_to_hms(start),
            "End (HH:MM:SS)":     seconds_to_hms(end),
            "Duration (HH:MM:SS)": seconds_to_hms(duration),
        }
        for aid, aname in attributes.items():
            row[aname] = av.get(aid, "")

        rows.append(row)

    # Sort by file then start time
    rows.sort(key=lambda r: (r["File"], r["Start (s)"]))

    # Write Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Annotations"

        if not rows:
            print("No annotations found.")
            return

        headers = list(rows[0].keys())
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, start=2):
            for col, key in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col, value=row[key])

        # Auto-size columns
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        wb.save(out_path)

    except ImportError:
        # Fallback to csv if openpyxl not installed
        import csv
        out_path = out_path.with_suffix(".csv")
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    print(f"Saved {len(rows)} annotations → {out_path}")


if __name__ == "__main__":
    main()
